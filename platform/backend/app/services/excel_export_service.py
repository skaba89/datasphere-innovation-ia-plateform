"""
Excel Export Service — Generate .xlsx reports from the platform data.
Uses openpyxl with styled headers, auto-column widths and freeze panes.
"""

from __future__ import annotations

import io
from datetime import datetime

from sqlalchemy.orm import Session

from app.models.agent import AgentAction, AgentAssignment, AgentProfile
from app.models.deliverable import Deliverable
from app.models.opportunity import Opportunity
from app.models.organization import Organization
from app.models.tender import Tender


# ── Styling helpers ──────────────────────────────────────────────────────────

def _make_workbook():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    return wb


def _style_header_row(ws, header_row: int = 1):
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(name="Calibri", bold=True, color="FACC15", size=11)

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, min_width: int = 12, max_width: int = 50):
    for col in ws.columns:
        lengths = []
        for cell in col:
            if cell.value:
                lengths.append(len(str(cell.value)))
        if lengths:
            ws.column_dimensions[col[0].column_letter].width = min(max(max(lengths) + 2, min_width), max_width)


def _zebra_rows(ws, start_row: int = 2, light_color: str = "F8FAFC", dark_color: str = "EFF6FF"):
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.styles import Font

    body_font = Font(name="Calibri", size=10)
    for i, row in enumerate(ws.iter_rows(min_row=start_row)):
        fill_color = light_color if i % 2 == 0 else dark_color
        fill = PatternFill("solid", fgColor=fill_color)
        for cell in row:
            cell.fill = fill
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=False)


def _bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Report 1 — Pipeline commercial ──────────────────────────────────────────

def export_pipeline(db: Session) -> bytes:
    wb = _make_workbook()
    ws = wb.active
    ws.title = "Pipeline"
    ws.freeze_panes = "A2"

    headers = [
        "ID", "Titre", "Organisation", "Pays", "Secteur",
        "Statut", "Priorité", "Valeur potentielle (€)",
        "Probabilité (%)", "Pipeline pondéré (€)",
        "Prochaine action", "Responsable", "Créé le",
    ]
    ws.append(headers)

    rows = (
        db.query(Opportunity, Organization.name.label("org_name"))
        .join(Organization, Organization.id == Opportunity.organization_id)
        .order_by(Opportunity.priority.desc(), Opportunity.created_at.desc())
        .all()
    )

    for opp, org_name in rows:
        val = float(opp.potential_value or 0)
        prob = opp.probability or 0
        ws.append([
            opp.id,
            opp.title,
            org_name,
            opp.country or "",
            opp.sector or "",
            opp.status,
            opp.priority,
            round(val, 2),
            prob,
            round(val * prob / 100, 2),
            opp.next_action or "",
            opp.owner_name or "",
            opp.created_at.strftime("%d/%m/%Y") if opp.created_at else "",
        ])

    _style_header_row(ws)
    _zebra_rows(ws)
    _auto_width(ws)
    return _bytes(wb)


# ── Report 2 — Appels d'offres ───────────────────────────────────────────────

def export_tenders(db: Session) -> bytes:
    wb = _make_workbook()
    ws = wb.active
    ws.title = "Appels d'offres"
    ws.freeze_panes = "A2"

    headers = [
        "ID", "Référence", "Titre", "Acheteur",
        "Décision Go/No-Go", "Score (/100)", "Statut",
        "Date limite remise", "Nb exigences", "Créé le",
    ]
    ws.append(headers)

    from app.models.tender import TenderRequirement
    from sqlalchemy import func

    req_counts = dict(
        db.query(TenderRequirement.tender_id, func.count(TenderRequirement.id))
        .group_by(TenderRequirement.tender_id)
        .all()
    )

    for t in db.query(Tender).order_by(Tender.created_at.desc()).all():
        ws.append([
            t.id,
            t.reference or "",
            t.title,
            t.buyer_name or "",
            t.go_no_go_decision or "Non qualifié",
            t.go_no_go_score or 0,
            t.status,
            t.submission_deadline.strftime("%d/%m/%Y") if t.submission_deadline else "",
            req_counts.get(t.id, 0),
            t.created_at.strftime("%d/%m/%Y") if t.created_at else "",
        ])

    _style_header_row(ws)
    _zebra_rows(ws)
    _auto_width(ws)
    return _bytes(wb)


# ── Report 3 — Actions agents ────────────────────────────────────────────────

def export_actions(db: Session) -> bytes:
    wb = _make_workbook()
    ws = wb.active
    ws.title = "Actions agents"
    ws.freeze_panes = "A2"

    headers = [
        "ID Action", "Type", "Titre", "Priorité", "Statut",
        "Approbation requise", "Approuvé par", "Exécuté le",
        "Profil agent", "Assignment #",
        "Résumé résultat",
    ]
    ws.append(headers)

    rows = (
        db.query(AgentAction, AgentProfile.name.label("agent_name"))
        .join(AgentAssignment, AgentAssignment.id == AgentAction.assignment_id)
        .join(AgentProfile, AgentProfile.id == AgentAssignment.agent_id)
        .order_by(AgentAction.created_at.desc())
        .all()
    )

    for action, agent_name in rows:
        ws.append([
            action.id,
            action.action_type,
            action.title,
            action.priority,
            action.status,
            "Oui" if action.requires_human_approval else "Non",
            action.approved_by or "",
            action.executed_at.strftime("%d/%m/%Y %H:%M") if action.executed_at else "",
            agent_name,
            action.assignment_id,
            (action.result_summary or "")[:200],
        ])

    _style_header_row(ws)
    _zebra_rows(ws)
    _auto_width(ws)
    return _bytes(wb)


# ── Report 4 — Livrables ─────────────────────────────────────────────────────

def export_deliverables(db: Session) -> bytes:
    wb = _make_workbook()
    ws = wb.active
    ws.title = "Livrables"
    ws.freeze_panes = "A2"

    headers = [
        "ID", "Titre", "Type", "Statut", "Version", "Langue",
        "Audience", "Généré par", "Révisé par", "Approuvé par",
        "Date approbation", "Résumé",
    ]
    ws.append(headers)

    for d in db.query(Deliverable).order_by(Deliverable.created_at.desc()).all():
        ws.append([
            d.id,
            d.title,
            d.deliverable_type,
            d.status,
            d.version,
            d.language,
            d.audience or "",
            d.generated_by or "",
            d.reviewed_by or "",
            d.approved_by or "",
            d.approved_at.strftime("%d/%m/%Y") if d.approved_at else "",
            (d.summary or "")[:200],
        ])

    _style_header_row(ws)
    _zebra_rows(ws)
    _auto_width(ws)
    return _bytes(wb)


# ── Report 5 — Rapport complet multi-onglets ─────────────────────────────────

def export_full_report(db: Session) -> bytes:
    from openpyxl import load_workbook

    wb = _make_workbook()
    ws_cover = wb.active
    ws_cover.title = "Rapport DataSphere"

    # Cover sheet
    from openpyxl.styles import Alignment, Font
    ws_cover["A1"] = "DataSphere Innovation IA Platform"
    ws_cover["A1"].font = Font(name="Calibri", bold=True, size=20, color="0F172A")
    ws_cover["A2"] = f"Rapport complet — {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} UTC"
    ws_cover["A2"].font = Font(name="Calibri", size=12, color="64748B")
    ws_cover.column_dimensions["A"].width = 60

    # Merge cover cell
    ws_cover.merge_cells("A1:J1")
    ws_cover["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_cover.row_dimensions[1].height = 40

    # Helper to copy a report into a new sheet
    def _add_sheet(name: str, source_bytes: bytes):
        src_wb = load_workbook(io.BytesIO(source_bytes))
        src_ws = src_wb.active
        new_ws = wb.create_sheet(title=name)
        for row in src_ws.iter_rows(values_only=True):
            new_ws.append(row)
        _style_header_row(new_ws)
        _zebra_rows(new_ws)
        _auto_width(new_ws)
        new_ws.freeze_panes = "A2"

    _add_sheet("Pipeline", export_pipeline(db))
    _add_sheet("Appels d'offres", export_tenders(db))
    _add_sheet("Actions agents", export_actions(db))
    _add_sheet("Livrables", export_deliverables(db))

    return _bytes(wb)
